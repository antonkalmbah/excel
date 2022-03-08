from openpyxl import Workbook, load_workbook

wb = Workbook()
# ws_2 = wb.create_sheet('Page_two')
# ws_3 = wb.create_sheet('Page_three')

# ws_3.title = 'Page_3'
# переименование ws_3

# ws_2.sheet_properties.tabColor = 'FF0000'
# изменяем цвет листа внизу в excel

# ws_2 = wb('Page_two')

# ws = wb.active
# использование уже имеющегося рабочего листа

# write_1 = ws_2['A1']
# ws_2['A1'] = 'Новая запись из Python'
#
# write_2 = ws_2['A2']
# ws_2['A2'] = 'Вторая новая запись в Python'
#
# printer = ws_2.cell(row=2, column=1, value=2)
# доступ к 1 ячейке
# printer = ws_2['A1':'A2']

# print(printer)

# for row in ws_2.values(min_row=1, max_col=3, max_row=3, values_only=True):
#     print(row)
#
#
# wb.save('Машины.xlsx')

# wd = load_workbook('Машины.xlsx')
# # загружаем нужную таблицу и не забываем load_workbook импортировать перед этим
# wb.template = True
# wb.save('Машины_копия.xlsx')

# wb = load_workbook(filename='Машины.xlsx', read_only=True)
# sheet = wb['Sheet']


# print(sheet['A1'].value)

# def create_table():
#     for row in sheet.rows:
#         for cell in row:
#             print(cell.value)
# создали функцию, которая выводит в консоль все данные из таблицы и дальше с помощью wb.save('new.xlsx') записывает их
# в нужный файл
